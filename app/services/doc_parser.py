"""文档解析服务：将 PDF/Word/TXT/Excel 文件解析为文本切片"""
from __future__ import annotations

import io
from typing import List

from loguru import logger
from langchain_text_splitters import RecursiveCharacterTextSplitter

_CHUNK_SIZE = 500
_CHUNK_OVERLAP = 80


def parse_file(content: bytes, filename: str) -> List[str]:
    """
    根据文件后缀选择解析方式，返回文本切片列表。
    支持: .pdf, .docx, .txt, .md, .xlsx, .csv
    """
    name_lower = filename.lower()
    logger.info("[DocParser] 开始解析: file={}, size={}KB", filename, len(content) // 1024)

    if name_lower.endswith(".pdf"):
        logger.info("[DocParser] 文件类型: PDF")
        text = _parse_pdf(content)
    elif name_lower.endswith(".docx"):
        logger.info("[DocParser] 文件类型: Word (docx)")
        text = _parse_docx(content)
    elif name_lower.endswith((".txt", ".md", ".csv")):
        logger.info("[DocParser] 文件类型: 纯文本 ({})", name_lower.rsplit(".", 1)[-1])
        text = content.decode("utf-8", errors="ignore")
    elif name_lower.endswith(".xlsx"):
        logger.info("[DocParser] 文件类型: Excel (xlsx)")
        text = _parse_xlsx(content)
    else:
        logger.warning("[DocParser] 不支持的文件类型: {}", filename)
        return []

    logger.info("[DocParser] 原文提取完成: {} 字符", len(text))

    if not text.strip():
        logger.warning("[DocParser] 文件内容为空: {}", filename)
        return []

    chunks = _split_text(text)
    logger.info("[DocParser] 切片完成: {} → {} 个切片 (chunk_size={}, overlap={})",
                filename, len(chunks), _CHUNK_SIZE, _CHUNK_OVERLAP)
    return chunks


def _parse_pdf(content: bytes) -> str:
    import pymupdf
    doc = pymupdf.open(stream=content, filetype="pdf")
    logger.info("[DocParser] PDF 共 {} 页", len(doc))
    pages = []
    for i, page in enumerate(doc):
        page_text = page.get_text()
        pages.append(page_text)
        if i < 3:
            logger.info("[DocParser] 第{}页: {}字符", i + 1, len(page_text))
    doc.close()
    return "\n".join(pages)


def _parse_docx(content: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    logger.info("[DocParser] Word 共 {} 个段落", len(paragraphs))
    return "\n".join(paragraphs)


def _parse_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"[Sheet: {sheet.title}]")
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells).strip()
            if line and line != "|":
                lines.append(line)
                row_count += 1
        logger.info("[DocParser] Sheet '{}': {} 行", sheet.title, row_count)
    wb.close()
    return "\n".join(lines)


def _split_text(text: str) -> List[str]:
    """
    递归字符分块：
    优先按段落/句子等自然边界切分，过长时再按窗口兜底。
    """
    text = text.strip()
    if not text:
        return []

    splitter = RecursiveCharacterTextSplitter(
        # 从粗到细的分隔符：段落 -> 行 -> 句子 -> 逗号/空格
        separators=["\n\n", "\n", "。", "！", "？", "；", "，", " "],
        chunk_size=_CHUNK_SIZE,
        chunk_overlap=_CHUNK_OVERLAP,
    )
    chunks = splitter.split_text(text)
    # 去掉空白块
    return [c.strip() for c in chunks if c.strip()]
